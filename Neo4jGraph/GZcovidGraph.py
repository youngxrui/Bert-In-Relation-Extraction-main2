from py2neo import Graph, Node, Relationship
from openpyxl import load_workbook
# Graph()中第一个为local host链接，auth为认证，包含 username 和 password
# graph = Graph('http://localhost:7474', auth = ('username', 'password'))
graph = Graph('http://localhost:7474', auth = ('neo4j', 'COVID'))
# a = Node("hero", name="Clint")  # Node(label, name)
# b = Node("hero", name="Natasha")
# ab = Relationship(a, "friend", b)
# print(a,b,ab)
# graph.create(ab)  # 创建节点和关系
# # graph.delete_all()
# c = Node('Person', name='Alice')
# d = Node('Person', name='Bob')
# e = Node('Person','hero', name='Bob')#Bob对应'Person'和'hero'两个标签
# r = Relationship(c, 'KNOWS', d)
# graph.create(e)
# print(c, d, r)

def ageGenderCreate2(file,sheet1,e1_id_column,e1_name_column,r_label_column,e2_column):
    '''
    :param file:excel
    :param sheet1: 第几个sheet
    :param e1_id_column:e1_id所在列
    :param e1_name_column:e1所在列
    :param r_label_column:r_label所在列
    :param e2_column:e2所在列，即属性值
    :return:
    '''
    wb1 = load_workbook(file)
    sheet1 = wb1[wb1.sheetnames[sheet1 - 1]]
    sheet1_max_row = sheet1.max_row
    idList=[]
    for i in range(2,sheet1_max_row+1):
        print("第{}行".format(i))
        if sheet1.cell(i, e1_id_column).value in idList:
            print("该id-{}节点已录入".format(sheet1.cell(i, e1_id_column).value))
        else:
            e1_id=sheet1.cell(i,e1_id_column).value
            e1_name = sheet1.cell(i, e1_name_column).value
            r_label_1 = sheet1.cell(i, r_label_column).value
            gender_=""
            age_=""
            if r_label_1 == "gender":
                gender_ = sheet1.cell(i, e2_column).value
            else:
                age_ = sheet1.cell(i, e2_column).value
            for j in range(i+1,i+3):
                if sheet1.cell(j,e1_id_column).value==e1_id:
                   r_label_2=sheet1.cell(j,r_label_column).value
                   if r_label_2=="gender":
                       gender_=sheet1.cell(j,e2_column).value
                   else:
                       age_ = sheet1.cell(j,e2_column).value
            print(e1_name,e1_id,gender_, age_)
            node=Node("infected",name=e1_name,id=e1_id,gender=gender_,age=age_)
            idList.append(e1_id)
            graph.create(node)
    wb1.close()
    filename = open('a.txt', 'w')
    for value in idList:
        filename.write(str(value))
        filename.write("\n")
    filename.close()
    return idList


# 建立无属性的其他实体
def createOtherEntity(file,sheet1,idList,e1_id_column,e1_name_column,e1_label_column,e2_id_column,e2_name_column,e2_label_column):
    '''

    '''
    wb1 = load_workbook(file)
    sheet1 = wb1[wb1.sheetnames[sheet1 - 1]]
    sheet1_max_row = sheet1.max_row
    for i in range(2,sheet1_max_row+1):
        print("第{}行".format(i))
        if sheet1.cell(i, e1_id_column).value in idList:
            print("id为{}的节点已创建".format(sheet1.cell(i, e1_id_column).value))
        else:
            e1_name=sheet1.cell(i, e1_name_column).value
            e1_id = sheet1.cell(i, e1_id_column).value
            e1_label=sheet1.cell(i, e1_label_column).value
            node = Node(e1_label, name=e1_name, id=e1_id)
            idList.append(e1_id)
            graph.create(node)

        if sheet1.cell(i, e2_id_column).value in idList:
            print("id为{}的节点已创建".format(sheet1.cell(i, e2_id_column).value))
        else:
            e2_name = sheet1.cell(i, e2_name_column).value
            e2_id = sheet1.cell(i, e2_id_column).value
            e2_label = sheet1.cell(i, e2_label_column).value
            node = Node(e2_label, name=e2_name, id=e2_id)
            idList.append(e2_id)
            graph.create(node)
    wb1.close()
    filename = open('b.txt', 'w')
    for value in idList:
        filename.write(str(value))
        filename.write("\n")
    filename.close()
    return

# 存入所有实体
def run():
    idList=ageGenderCreate2('entity_relation2.xlsx',1,11,2,4,5)
    createOtherEntity('entity_relation2.xlsx', 2, idList, 13, 2, 3, 15, 6,7)

# run()

def createRelationTest():
    r="r:" + "contain" + "{" + "name"+":"+"'"+"包含"+"'"+"}"
    start_node="P20201118003"
    end_node="P20201118008"
    # createRelationCypher="MATCH (start_node),(end_node) where start_node.id={} and end_node.id={} create  (start_node)-[{}]->(end_node)".format(str(start_node),str(end_node),r)
    createRelationCypher = "MATCH (start_node),(end_node) where start_node.id=" + "'" + start_node + "'" + " and end_node.id=" + "'" + end_node + "'" + " create  (start_node)-["+r+"]->(end_node)"
    graph.run(cypher=createRelationCypher).data()
    return
    # p1=
    # c = Node('Person', name='Alice')
    # d = Node('Person', name='Bob')
    # e = Node('Person','hero', name='Bob')#Bob对应'Person'和'hero'两个标签
    # r = Relationship(c, 'KNOWS', d)
    # graph.create(e)

# createRelationTest()

def createRelation(file,sheet1,e1_id_column,e2_id_column,r_label_column):
    '''

    :param file: excel文件
    :param sheet1: sheet
    :param e1_id_column:头实体id所在列
    :param e2_id_column:尾实体id所在列
    :param r_label_column:关系label所在列
    :return:
    '''
    wb1 = load_workbook(file)
    sheet1 = wb1[wb1.sheetnames[sheet1 - 1]]
    sheet1_max_row = sheet1.max_row
    relationList=[]
    for i in range(2,sheet1_max_row+1):
        if i%10==0:
            print("第{}行,总共有{}行".format(i,sheet1_max_row))
        try:
            start_node=sheet1.cell(i,e1_id_column).value
            end_node=sheet1.cell(i,e2_id_column).value
            r="r:" + sheet1.cell(i,r_label_column).value
            ere=start_node+"-"+r+"-"+end_node
            if ere in relationList:
                print("实体关系{}已建立".format(ere))
            else:
                relationList.append(ere)
            # data = graph.data('MATCH (id:Person) return p')
                createRelationCypher = "MATCH (start_node),(end_node) where start_node.id=" + "'" + start_node + "'" + " and end_node.id=" + "'" + end_node + "'" + " create  (start_node)-[" + r + "]->(end_node)"
                graph.run(cypher=createRelationCypher).data()
        except:
            print("第{}条关系创建失败".format(str(i)))
    filename = open('ere.txt', 'w')
    for value in relationList:
        filename.write(str(value))
        filename.write("\n")
    filename.close()
    return
# createRelation('entity_relation2.xlsx', 2,13,15,5)
# createRelation('entity_relation3.xlsx', 5,13,15,5)#重新建立contain关系



# 添加坐标属性
def addXY(file,sheet1,x_column,y_column,e2_name_column):
    wb1 = load_workbook(file)
    sheet1 = wb1[wb1.sheetnames[sheet1 - 1]]
    sheet1_max_row = sheet1.max_row
    # relationList = []
    for i in range(2, sheet1_max_row+1):
        print("第{}行,总共有{}行".format(i, sheet1_max_row))
        e=sheet1.cell(i,e2_name_column).value
        x=sheet1.cell(i,x_column).value
        y = sheet1.cell(i, y_column).value
        try:
            createRelationCypher = "MATCH (n) where n.name=" + "'" + e + "'" + " SET n.x=" + "'"+str(x)+"', "+"n.y="+"'"+str(y)+"'"
            graph.run(cypher=createRelationCypher).data()
            setLabelPOI="MATCH (n) where n.name=" + "'" + e + "' SET n:POI RETURN n"
            graph.run(cypher=setLabelPOI).data()
        except:
            print("无此节点")
    return

# addXY(r"C:\Users\lenovo\Desktop\小论文\hanLP\Chinese_NRE-master\pythonProject\yqData\ere核查-时间统一-地点筛选-地点坐标获取 - 坐标转换 - 距离获取 - 部分重编码6.xlsx",8,23,24,4)


# 创建空间关系（距离和方位）
def createGeoRelation(file,sheet1,e1_name_column,e2_name_column,dist_value_column,dist_level_column,direction_column):

    wb1 = load_workbook(file)
    sheet1 = wb1[wb1.sheetnames[sheet1 - 1]]
    sheet1_max_row = sheet1.max_row
    relationList=[]
    for i in range(2,sheet1_max_row+1):
        if i%10==0:
            print("第{}行,总共有{}行".format(i,sheet1_max_row))
        try:
            start_node=sheet1.cell(i,e1_name_column).value
            end_node=sheet1.cell(i,e2_name_column).value
            spatial_r="r:spatial{distance:" + "'" + sheet1.cell(i,dist_value_column).value + "'," + "dist_level:" + "'" + sheet1.cell(i,dist_level_column).value + "'," + "direction:"+"'"+sheet1.cell(i,direction_column).value+"'}"
            ere=start_node+"-"+spatial_r+"-"+end_node
            if ere in relationList:
                print("实体关系{}已建立".format(ere))
            else:
                relationList.append(ere)
            # data = graph.data('MATCH (id:Person) return p')
                createRelationCypher = "MATCH (start_node),(end_node) where start_node.name=" + "'" + start_node + "'" + " and end_node.name=" + "'" + end_node + "'" + " create  (start_node)-[" + spatial_r + "]->(end_node)"
                graph.run(cypher=createRelationCypher).data()
        except:
            print("第{}条关系创建失败".format(str(i)))
    # filename = open('ereSpatial.txt', 'w')
    # for value in relationList:
    #     filename.write(str(value))
    #     filename.write("\n")
    # filename.close()
    return
# createGeoRelation('entity_relation3.xlsx',4,1,7,5,6,12)

# 创建病例间关系
def createPRelation(file,sheet1,e1_id_column,e2_id_column,r_name_column):

    wb1 = load_workbook(file)
    sheet1 = wb1[wb1.sheetnames[sheet1 - 1]]
    sheet1_max_row = sheet1.max_row
    relationList=[]
    for i in range(2,sheet1_max_row+1):
        if i%10==0:
            print("第{}行,总共有{}行".format(i,sheet1_max_row))
        try:
            start_node=sheet1.cell(i,e1_id_column).value
            end_node=sheet1.cell(i,e2_id_column).value
            PRelation_r="r:PRelation{name:" + "'" + sheet1.cell(i,r_name_column).value +"'}"
            ere=start_node+"-"+PRelation_r+"-"+end_node
            if ere in relationList:
                print("实体关系{}已建立".format(ere))
            else:
                relationList.append(ere)
            # data = graph.data('MATCH (id:Person) return p')
                createRelationCypher = "MATCH (start_node),(end_node) where start_node.id=" + "'" + start_node + "'" + " and end_node.id=" + "'" + end_node + "'" + " create  (start_node)-[" + PRelation_r + "]->(end_node)"
                graph.run(cypher=createRelationCypher).data()
        except:
            print("第{}条关系创建失败".format(str(i)))
    filename = open('PRelation.txt', 'w')
    for value in relationList:
        filename.write(str(value))
        filename.write("\n")
    filename.close()
    return
# createPRelation('entity_relation4.xlsx',6,13,15,6)

# 将时间的年月日替换
def renameDateName(file,sheet1,old_name_column,new_name_column):

    wb1 = load_workbook(file)
    sheet1 = wb1[wb1.sheetnames[sheet1 - 1]]
    sheet1_max_row = sheet1.max_row
    # relationList=[]
    for i in range(2,sheet1_max_row+1):
        if i%10==0:
            print("第{}行,总共有{}行".format(i,sheet1_max_row))
        try:
            old_name=sheet1.cell(i,old_name_column).value
            new_name=sheet1.cell(i,new_name_column).value

            # data = graph.data('MATCH (id:Person) return p')
            renameCypher = "MATCH (d) where d.name=" + "'" + old_name + "'" + " SET d.English_name=" + "'" + new_name + "'"
            graph.run(cypher=renameCypher).data()
        except:
            print("第{}条重命名失败".format(str(i)))
    filename = open('PRelation.txt', 'w')
    # for value in relationList:
    #     filename.write(str(value))
    #     filename.write("\n")
    # filename.close()
    return
# renameDateName('entity_relation5.xlsx',3,1,3)
renameDateName('entity_relation5.xlsx',4,1,2)