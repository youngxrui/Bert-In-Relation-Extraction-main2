import numpy as np
from sklearn.preprocessing import MinMaxScaler
import time
import copy
import torch.nn as nn
import torch.nn.functional as F
import torch.optim as optim
import torch.nn.functional
from torch.utils.data import Dataset, DataLoader
import warnings
import torch
import time
import argparse
import json
import os
from transformers import BertTokenizer
from model import BERT_Classifier
from openpyxl import load_workbook


#from transformers import BertPreTrainedModel
import random
def setup_seed(seed):
     torch.manual_seed(seed)#设置CPU生成随机数的种子，方便下次复现实验结果
     torch.cuda.manual_seed_all(seed)#表示在GPU中设置生成随机数的种子
     np.random.seed(seed)
     random.seed(seed)
warnings.filterwarnings("ignore")
setup_seed(44)

from transformers import BertModel

from loader import map_id_rel

rel2id, id2rel = map_id_rel()

print(len(rel2id))
print(id2rel)

USE_CUDA = torch.cuda.is_available()

def test2(net_path,text_list,ent1_list,ent2_list,show_result=False,result=[]):
    '''
    :param net_path: 训练的模型文件路径
    :param text_list:文本list
    :param ent1_list:头实体list
    :param ent2_list:尾实体list
    :param result:实际的关系
    :param show_result:是否将预测错误的关系打印出来
    :return:
    '''

    max_length=256
    net=torch.load(net_path)

    net.eval()
    # 在模型中，我们通常会加上Dropout层和batch
    # normalization层，在模型预测阶段，我们需要将这些层设置到预测模式，model.eval()
    # 就是帮我们一键搞定的，如果在预测的时候忘记使用model.eval()，会导致不一致的预测结果。
    # eval（）时，pytorch会自动把BN和DropOut固定住，不会取平均，而是用训练好的值
    if USE_CUDA:
        net = net.cuda()
    rel_list = []
    correct=0
    total=0
    with torch.no_grad():
        print("关系列表长度:"+str(len(result)))
        if len(result)>1:#传入有正确关系的参数，计算正确率
            for text,ent1,ent2,label in zip(text_list,ent1_list,ent2_list,result):
                # zip() 函数用于将可迭代的对象作为参数，将对象中对应的元素打包成一个个元组，然后返回由这些元组组成的列表。
                # 如果各个迭代器的元素个数不一致，则返回列表长度与最短的对象相同，利用 * 号操作符，可以将元组解压为列表
                sent = ent1 + ent2+ text
                tokenizer = BertTokenizer.from_pretrained('bert-base-chinese')
                # 导入了bert-base-chinese预训练模型
                indexed_tokens = tokenizer.encode(sent, add_special_tokens=True)
                # 对我们的句子进行token to id的转化
                # encode方法可以一步到位地生成对应模型的输入
                # add_special_tokens: bool = True  将句子转化成对应模型的输入形式，默认开启
                # max_length 设置最大长度，如果不设置的话原模型设置的最大长度是512，此时，如果句子长度超过512会报错：
                avai_len = len(indexed_tokens)
                while len(indexed_tokens) < max_length:
                    indexed_tokens.append(0)  # 0 is id for [PAD]
                indexed_tokens = indexed_tokens[: max_length]
                indexed_tokens = torch.tensor(indexed_tokens).long().unsqueeze(0)  # (1, L)升维

                # Attention mask
                att_mask = torch.zeros(indexed_tokens.size()).long()  # (1, L)
                #函数torch.zeros()返回一个由标量值0填充的张量，其形状由变量参数size定义。
                att_mask[0, :avai_len] = 1
                if USE_CUDA:
                    # 转为GPU张量
                    indexed_tokens = indexed_tokens.cuda()
                    att_mask = att_mask.cuda()

                outputs = net(indexed_tokens, att_mask)
                # print(y)
                logits = outputs[1]#最终的全连接层的输出
                _, predicted = torch.max(logits.data, 1)
                # torch.max
                # 输入
                # input是softmax函数输出的一个tensor
                # dim是max函数索引的维度0 / 1，0 是每列的最大值，1是每行的最大值
                # 输出
                # 函数会返回两个tensor，第一个tensor是每行的最大值；第二个tensor是每行最大值的索引。
                result = predicted.cpu().numpy().tolist()[0]

                if show_result & (id2rel[result]!=label):
                    print("Source Text: ",text)
                    print("Entity1: ",ent1," Entity2: ",ent2," Predict Relation: ",id2rel[result]," True Relation: ",label)
                if id2rel[result]==label:
                    correct+=1
                total+=1
                #print('\n')
                rel_list.append(id2rel[result])
            print(correct," ",total," ",correct/total)
            # print(outputs)
            return rel_list,correct,total
        else:#未传入有正确关系的参数，即预测，不计算正确率，因为没有正确答案也无法计算正确率
            for text,ent1,ent2 in zip(text_list,ent1_list,ent2_list):
                # zip() 函数用于将可迭代的对象作为参数，将对象中对应的元素打包成一个个元组，然后返回由这些元组组成的列表。
                # 如果各个迭代器的元素个数不一致，则返回列表长度与最短的对象相同，利用 * 号操作符，可以将元组解压为列表
                sent = ent1 + ent2+ text
                tokenizer = BertTokenizer.from_pretrained('bert-base-chinese')
                # 导入了bert-base-chinese预训练模型
                indexed_tokens = tokenizer.encode(sent, add_special_tokens=True)
                # 对我们的句子进行token to id的转化
                # encode方法可以一步到位地生成对应模型的输入
                # add_special_tokens: bool = True  将句子转化成对应模型的输入形式，默认开启
                # max_length 设置最大长度，如果不设置的话原模型设置的最大长度是512，此时，如果句子长度超过512会报错：
                avai_len = len(indexed_tokens)
                while len(indexed_tokens) < max_length:
                    indexed_tokens.append(0)  # 0 is id for [PAD]
                indexed_tokens = indexed_tokens[: max_length]
                indexed_tokens = torch.tensor(indexed_tokens).long().unsqueeze(0)  # (1, L)升维

                # Attention mask
                att_mask = torch.zeros(indexed_tokens.size()).long()  # (1, L)
                #函数torch.zeros()返回一个由标量值0填充的张量，其形状由变量参数size定义。
                att_mask[0, :avai_len] = 1
                if USE_CUDA:
                    # 转为GPU张量
                    indexed_tokens = indexed_tokens.cuda()
                    att_mask = att_mask.cuda()

                outputs = net(indexed_tokens, att_mask)
                # print(y)
                logits = outputs[1]#最终的全连接层的输出
                _, predicted = torch.max(logits.data, 1)
                # torch.max
                # 输入
                # input是softmax函数输出的一个tensor
                # dim是max函数索引的维度0 / 1，0 是每列的最大值，1是每行的最大值
                # 输出
                # 函数会返回两个tensor，第一个tensor是每行的最大值；第二个tensor是每行最大值的索引。
                result = predicted.cpu().numpy().tolist()[0]
                print("Source Text: ", text)
                print("Entity1: ", ent1, " Entity2: ", ent2, " Predict Relation: ", id2rel[result])
                # 实体关系写入表格
                # wb1 = load_workbook(excelfile)
                # sheet1 = wb1[wb1.sheetnames[7]]
                # max_row = sheet1.max_row
                # max_column = sheet1.max_column

                #print('\n')
            # print(outputs)
            return id2rel[result]



from random import choice

def demo_output():
    text_list=[]
    ent1=[]
    ent2=[]
    result=[]
    total_num=300#输出50个提取示例
    # with open("train.json", 'r', encoding='utf-8') as load_f:
    with open("data\核查train.json", 'r', encoding='utf-8') as load_f:
        lines=load_f.readlines()       
        while total_num>0:
            line=choice(lines)#?????????
            dic = json.loads(line)
            text_list.append(dic['text'])
            ent1.append(dic['ent1'])
            ent2.append(dic['ent2'])
            result.append(dic['rel'])
            total_num-=1
            if total_num<0:
                break
    test2('0.9648148148148148.pth', text_list, ent1, ent2,True, result)
    # test('0.9537394662921348.pth', text_list, ent1, ent2, result, True)


# 计算每一个类别的正确率
def caculate_acc():
    correct0=0
    total0=0
    for i in range(len(rel2id)):
        temp_rel = id2rel[i]
        text_list=[]
        ent1=[]
        ent2=[]
        result=[]
        # with open("dev.json", 'r', encoding='utf-8') as load_f:
        with open("data\核查test.json", 'r', encoding='utf-8') as load_f:
            lines = load_f.readlines()
            for line in lines:
                line=choice(lines)#方法返回一个列表，元组或字符串的随机项??????????
                dic = json.loads(line)#json.loads 用于解码 JSON 数据。该函数返回 Python 字段的数据类型
                if dic['rel']==temp_rel:
                    text_list.append(dic['text'])
                    ent1.append(dic['ent1'])
                    ent2.append(dic['ent2'])
                    result.append(dic['rel'])
                if len(text_list)==100:#每个自定义的关系在测试集中选取100条数据
                    break
        if len(text_list) == 0:
            print("No sample: ", temp_rel)
        else:
            # test('0.8542471890372453.pth', text_list, ent1, ent2, result,True)

            rel_list, correct, total = test2('0.992948717948718.pth', text_list, ent1, ent2, True, result)
            correct0+=correct
            total0+=total
    percent=correct0/total0
    print(str(total0)+"条数据准确率为："+str(percent))
            # test('0.9537394662921348.pth', text_list, ent1, ent2, result)
# net = torch.load('0.3827081869290232.pth', map_location=torch.device('cpu'))
# print(type(net))  # 类型是 dict
# print(len(net))  # 长度为 4，即存在四个 key-value 键值对

#关系抽取
def predict():
    # correct0=0
    # total0=0
    for i in range(len(rel2id)):
        temp_rel = id2rel[i]
        text_list=[]
        ent1=[]
        ent2=[]
        # result=[]

    with open("data\predict_data.json", 'r', encoding='utf-8') as load_f:
        lines = load_f.readlines()
        for line in lines:
            line=choice(lines)#方法返回一个列表，元组或字符串的随机项??????????
            dic = json.loads(line)#json.loads 用于解码 JSON 数据。该函数返回 Python 字段的数据类型

            text_list.append(dic['text'])
            ent1.append(dic['ent1'])
            ent2.append(dic['ent2'])
            # result.append(dic['rel'])
        a=test2('0.992948717948718.pth', text_list, ent1, ent2,True)
    # text_list = [
    #     "2021年12月16日广州市新冠肺炎疫情情况,新增1例境外输入关联确诊病例情况：女，70岁，中国籍，退休人员，居住广州越秀区华乐街天胜村65号。与12月13日境外输入外市返穗确诊病例的居所为同一栋楼，作为该病例的涉疫风险人员，于12月13日当天闭环转运至集中隔离场所进行医学观察，发病前已落实集中隔离管控。13日、14日、15日核酸检测结果均为阴性，16日凌晨，核酸检测结果初筛阳性，市疾控中心复核阳性，即闭环转运至广州医科大学附属市八医院隔离医学观察。经进一步检查和专家会诊，诊断为新冠肺炎确诊病例。",
    #     "2021年12月16日广州市新冠肺炎疫情情况,新增1例境外输入关联确诊病例情况：女，70岁，中国籍，退休人员，居住广州越秀区华乐街天胜村65号。与12月13日境外输入外市返穗确诊病例的居所为同一栋楼，作为该病例的涉疫风险人员，于12月13日当天闭环转运至集中隔离场所进行医学观察，发病前已落实集中隔离管控。13日、14日、15日核酸检测结果均为阴性，16日凌晨，核酸检测结果初筛阳性，市疾控中心复核阳性，即闭环转运至广州医科大学附属市八医院隔离医学观察。经进一步检查和专家会诊，诊断为新冠肺炎确诊病例。",
    #     ]
    # ent1 = ['确诊病例', '确诊病例']
    # ent2 = ['退休人员', '中国']
    # result = test2('0.992948717948718.pth', text_list, ent1, ent2, True)
    return


def extractRelation():
    # correct0=0
    # total0=0
    for i in range(len(rel2id)):
        temp_rel = id2rel[i]
        # text_list=[]
        # ent1=[]
        # ent2=[]
        # result=[]
#读取含待抽取关系的实体对表格，思路：逐句读取将每一句存在的所有实体对存入列表并将关系写入表格
    wb = load_workbook("data\广州市疫情数据除20210228-20220528分词-关系抽取数据-汇总 - 副本.xlsx",read_only=False)
    sheet1 = wb[wb.sheetnames[6]]
    maxRow=sheet1.max_row
    maxColumn=sheet1.max_column
    print("当前sheet名为：{}，共有：{}行，最大列{}".format(wb.sheetnames[6],maxRow,maxColumn))
    start_row=4 #含待识别实体对的文本所在起始行，后面一句+3

    while start_row<maxRow+1:
    # while start_row < 100:
        start_column = 11  # 该句第一个实体对的头实体，后面实体对的头实体+4
        print("当前行数{}，当前进度：{}".format(start_row,start_row/maxColumn))
        if sheet1.cell(start_row,5).value=='1':#判断该句是否有带提取关系的实体对
           while start_column<maxColumn-1:#maxColumn是该句最后一个实体对的尾实体
                text_list = []
                ent1 = []
                ent2 = []
                text_list.append(sheet1.cell(start_row,8).value)
                if (sheet1.cell(start_row,start_column).value != None) & (sheet1.cell(start_row,start_column+2).value != None):
                    ent1.append(sheet1.cell(start_row,start_column).value)
                    ent2.append(sheet1.cell(start_row,start_column+2).value)
                    relation = test2('0.992948717948718.pth', text_list, ent1, ent2, True)
                    sheet1.cell(start_row, start_column + 1).value=relation#将预测的关系写入表中
                    start_column = start_column + 4
                    # print("###############")
                else:
                    # print("$$$$$$$$$")
                    start_column = start_column + 4

           start_row = start_row + 3
        else:
            start_row = start_row + 3

    wb.save("data\广州市疫情数据除20210228-20220528分词-关系抽取数据-汇总 - 副本.xlsx")
    wb.close()
    return




# demo_output()
# caculate_acc()
# predict()
extractRelation()